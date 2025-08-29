import io
import re
import os
import jieba
import pytesseract
from PIL import Image
from tqdm import tqdm
import pandas as pd
from sentence_transformers import util

import openpyxl
from openpyxl.utils import get_column_letter

import public_function as pf

class ExcelProcessor:
    def __init__(self, embedding_model, summary_generator):
        self.embedding_model = embedding_model
        self.summary_generator = summary_generator

    def parse_excel(self, file_path):
        # df = pd.read_excel(file_path, sheet_name=None)
        # print(df.keys())
        # print(df.values())
        # 用 openpyxl 读取 Excel（保留合并单元格信息）

        # 存储所有结果
        all_results = {}
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            if sheet_name == '节点说明_20240227':
                sheet = wb[sheet_name]
                # print(f"正在处理 Sheet：{sheet_name}")

                # 获取所有合并单元格范围
                merged_ranges = sheet.merged_cells.ranges

                result = []

                row = 1
                while row <= sheet.max_row:
                    merged_block_size = 1
                    header_value = sheet.cell(row=row, column=1).value

                    # 检查是否是合并单元格起始行
                    for merged_range in merged_ranges:
                        if merged_range.start_cell.row == row and merged_range.start_cell.column == 1:
                            merged_block_size = merged_range.max_row - merged_range.min_row + 1
                            break

                    # 从第二列开始按行提取内容，合并成一个列表元素
                    block_content = []
                    for i in range(row, row + merged_block_size):
                        row_values = []
                        for j in range(2, sheet.max_column + 1):  # 第2列开始
                            cell_value = sheet.cell(row=i, column=j).value
                            if cell_value is not None:
                                row_values.append(str(cell_value).strip())
                        if row_values:
                            row_values.insert(0, header_value)
                            full_row_text = row_values.copy()
                            block_content.append(full_row_text)

                    # 合并单元格内容
                    blocks_group = ['|'.join(block) for block in block_content]

                    result.append('\n'.join(blocks_group))
                    row += merged_block_size  # 移动到下一个单元格块

                all_results[sheet_name] = '\n'.join([item for item in result if item])
            else:
                sheet = wb[sheet_name]
                # 获取所有合并单元格范围
                merged_ranges = sheet.merged_cells.ranges

                result = []

                row = 1
                while row <= sheet.max_row:
                    merged_block_size = 1
                    header_value = sheet.cell(row=row, column=1).value

                    # 检查第一列是否是合并单元格
                    for merged_range in merged_ranges:
                        if merged_range.start_cell.row == row and merged_range.start_cell.column == 1:
                            merged_block_size = merged_range.max_row - merged_range.min_row + 1
                            break

                    # 检查后续列是否也是合并单元格
                    is_full_block = True
                    for j in range(2, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row, column=j).value
                        if cell_value is not None:
                            is_full_block = False
                            break

                    if is_full_block:
                        # 如果后续列也是合并单元格
                        block_content = []
                        for merged_range in merged_ranges:
                            if merged_range.start_cell.row == row and merged_range.start_cell.column > 1:
                                block_size = merged_range.max_row - merged_range.min_row + 1
                                col_values = []
                                for i in range(merged_range.min_row, merged_range.max_row + 1):
                                    col_values.append(sheet.cell(row=i, column=merged_range.start_cell.column).value)
                                col_values.insert(0, header_value)
                                block_content.append('|'.join(map(str, col_values)))
                        result.append('\n'.join(block_content))
                    else:
                        # 如果后续列是独立内容
                        block_content = []
                        for j in range(2, sheet.max_column + 1):
                            col_values = []
                            for i in range(row, row + merged_block_size):
                                cell_value = sheet.cell(row=i, column=j).value
                                if cell_value is not None:
                                    col_values.append(str(cell_value).strip())
                            if col_values:
                                block_content.append('|'.join(col_values))
                        block_content.insert(0, str(header_value))
                        result.append('\n'.join(block_content))

                    row += merged_block_size  # 移动到下一个单元格块

                all_results[sheet_name] = '\n'.join([item for item in result if item])

        excel_content = []
        for key, value in all_results.items():
            if key and value:
                excel_content.append('\n'.join([key, value]))

        return excel_content

    def tokenize_text_jieba(self, text):
        """
        使用 jieba 对中英文混合文本进行分词：
        - 中文用 jieba
        - 英文/数字用正则保留原格式
        """
        # jieba 对整体做一次初步分词
        words = jieba.lcut(text)
        tokens = []
        for word in words:
            # 英文、数字、IP地址、文件名等直接识别
            if re.match(r'^[a-zA-Z0-9_\-./\\]+$', word):
                tokens.append(word)
            else:
                # 否则按中文词添加（可能是单字或词组）
                for char in word:
                    if '\u4e00' <= char <= '\u9fff':
                        tokens.append(char)
                    else:
                        tokens.append(char)

        return tokens

    def slide_split(self, text, max_len=500, overlap=50):
        """基于 jieba 中文分词 + 英文正则 分块"""
        tokens = self.tokenize_text_jieba(text)
        chunks = []
        start = 0
        
        while start < len(tokens):
            end = min(start + max_len, len(tokens))
            chunk_tokens = tokens[start:end]
            chunks.append(''.join(chunk_tokens))
            if end == len(tokens):
                break
            start += (max_len - overlap)

        return chunks
        
    def cluster_by_similarity(self, chunks, threshold=0.65, max_group_size=3):
        """基于相似度合并为章节，并生成摘要"""
        merged_sections = []
        current_group = []

        with tqdm(total=len(chunks), desc='相似度计算') as pbar:
            i = 0
            while i < len(chunks):
                current = chunks[i]
                prev_start = max(0, i - 2)
                prev_neighbors = chunks[prev_start: i]
                next_end = min(len(chunks), i + 3)
                next_neighbors = chunks[i + 1: next_end]

                # 前两块与后两块作为当前块的邻居
                neighbors = prev_neighbors + next_neighbors

                if neighbors:
                    emb_current = self.embedding_model.encode(current, normalize_embeddings=True)
                    emb_neighbors = self.embedding_model.encode(neighbors, normalize_embeddings=True)
                    sim_scores = self.embedding_model.similarity(emb_current, emb_neighbors)
                    # sim_scores = util.cos_sim(emb_current, emb_neighbors)
                    avg_sim = sim_scores.mean().item()

                    if avg_sim > threshold and len(current_group) < max_group_size:
                        current_group.append(current)
                    else:
                        # 若当前已有聚合块，先处理它
                        if current_group:
                            full_text = '。'.join(current_group)
                            summary = self.summary_generator.gen_summary(full_text)
                            merged_sections.append({
                                'summary': summary,
                                'chunks': current_group.copy()
                            })
                            current_group.clear()
                        current_group.append(current)
                else:
                    current_group.append(current)

                pbar.update(1)
                i += 1

        # 处理最后一组
        if current_group:
            full_text = '。'.join(current_group)
            summary = self.summary_generator.gen_summary(full_text)
            merged_sections.append({
                'summary': summary,
                'chunks': current_group.copy()
            })

        return merged_sections
    
    def read_excel(self, file_path: str, dataset) -> list:
        """入口：读取 excel 文件，处理后返回 [summary + chunks] 列表"""
        # 创建文件保存路径
        dataset_path = f'preprocess_data_files/{dataset}/'
        os.makedirs(dataset_path, exist_ok=True)

        content_save_directory = dataset_path + 'content/'
        summary_save_directory = dataset_path + 'summary_500/'
        os.makedirs(content_save_directory, exist_ok=True)
        os.makedirs(summary_save_directory, exist_ok=True)

        file_name = file_path.split('/')[-1:][0].replace('.xlsx', '.txt')
        content_save_path = os.path.join(content_save_directory, file_name)
        summary_save_path = os.path.join(summary_save_directory, file_name)

        # if os.path.exists(content_save_path) and os.path.exists(summary_save_path):
        #     return pf.read_data_from_txt(summary_save_path)

        # 1. 读取文档，提取文本、图片、表格
        if os.path.exists(content_save_path):
            excel_contents = pf.read_data_from_txt(content_save_path)
        else:
            excel_contents = self.parse_excel(file_path)
            pf.write_data_to_txt(excel_contents, content_save_path)

        # 2. 滑动分块
        text = '\n'.join(excel_contents)
        chunks = self.slide_split(text)

        # 3. 相似度合并
        clustered_sections = self.cluster_by_similarity(chunks)
        pf.write_data_to_txt(clustered_sections, summary_save_path)

        return clustered_sections

